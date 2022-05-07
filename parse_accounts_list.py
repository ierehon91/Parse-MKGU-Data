import openpyxl


def get_mkgu_accounts_list():
    wb = openpyxl.load_workbook(r'mkgu_accounts_list.xlsx')
    sheet = wb['mkgu_accounts_list']

    mkgu_accouts_list = []
    for i in range(2, sheet.max_row + 1):
        login = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        if login is not None:
            mkgu_accouts_list.append({'login': login, 'password': password})
        else:
            break
    return mkgu_accouts_list


def get_mkgu_accounts_list_txt():
    accounts_file = open('accounts.txt', 'r')
    accounts_string = accounts_file.read().split('\n')
    print(accounts_string)
    accounts = []
    for account in accounts_string:
        login_password = account.split(' ')
        login = login_password[0]
        password = login_password[1]
        accounts.append({'login': login, 'password': password})
    return accounts


def print_counts_accounts(mkgu_accouts_list):
    message = f'Найдено {len(mkgu_accouts_list)} учетных записей МКГУ.'
    print(message)
